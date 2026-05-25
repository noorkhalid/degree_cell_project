from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from degree.permissions import role_required
from accounts.models import UserProfile
from .forms import BankForm, CampusExcelUploadForm, CampusForm, DepartmentExcelUploadForm, DepartmentForm, ProgramExcelUploadForm, ProgramForm, CourierCompanyForm
from .models import Bank, Campus, Department, Program, CourierCompany

CAMPUS_TEMPLATE_HEADERS = ['Name', 'Active']
DEPARTMENT_TEMPLATE_HEADERS = ['Campus', 'Department', 'Active']
PROGRAM_TEMPLATE_HEADERS = ['Name', 'Level', 'Active']


def _delete_setup_item(request, model_class, object_name, redirect_url):
    item = get_object_or_404(model_class, pk=request.resolver_match.kwargs['pk'])
    if request.method == 'POST':
        try:
            item.delete()
            messages.success(request, f'{object_name} deleted successfully.')
        except ProtectedError:
            messages.error(request, f'This {object_name.lower()} cannot be deleted because it is already used in records. You can edit it and set Active = False instead.')
        return redirect(redirect_url)
    return render(request, 'academics/delete_confirm.html', {'item': item, 'title': f'Delete {object_name}', 'object_name': object_name, 'cancel_url': redirect_url})


def _normalize_program_level(value):
    text = str(value or '').strip()
    if not text:
        return None
    upper_text = text.upper()
    for code, label in Program.Level.choices:
        if upper_text == code.upper() or upper_text == label.upper():
            return code
    return None


def _normalize_active(value):
    if value is None or str(value).strip() == '':
        return True
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {'true', 'yes', 'y', '1', 'active'}:
        return True
    if text in {'false', 'no', 'n', '0', 'inactive'}:
        return False
    return True


def _xlsx_response(workbook, filename):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


@role_required(UserProfile.Role.ADMIN)
def campus_template_download(request):
    wb = Workbook(); ws = wb.active; ws.title = 'Campuses'
    ws.append(CAMPUS_TEMPLATE_HEADERS); ws.append(['Main Campus', 'TRUE']); ws.append(['Sub Campus', 'TRUE'])
    ws['D1'] = 'Active values'; ws['D2'] = 'TRUE'; ws['D3'] = 'FALSE'
    for col, width in {'A': 38, 'B': 14, 'D': 16}.items(): ws.column_dimensions[col].width = width
    for cell in ws[1]: cell.font = Font(bold=True)
    return _xlsx_response(wb, 'campus_upload_template.xlsx')


@role_required(UserProfile.Role.ADMIN)
def campus_upload_excel(request):
    if request.method != 'POST': return redirect('academics:campuses')
    form = CampusExcelUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, 'Please select a valid Excel file.'); return redirect('academics:campuses')
    try:
        sheet = load_workbook(form.cleaned_data['excel_file'], data_only=True).active
    except Exception:
        messages.error(request, 'The uploaded file could not be read. Please use the Excel template.'); return redirect('academics:campuses')
    created = updated = skipped = 0; errors = []
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = str(row[0] or '').strip() if len(row) > 0 else ''; active_value = row[1] if len(row) > 1 else True
        if not name and not active_value: continue
        if not name: skipped += 1; errors.append(f'Row {index}: campus name is missing.'); continue
        _, was_created = Campus.objects.update_or_create(name=name, defaults={'is_active': _normalize_active(active_value)})
        created += int(was_created); updated += int(not was_created)
    messages.success(request, f'Excel upload complete. Created: {created}, Updated: {updated}, Skipped: {skipped}.')
    if errors: messages.warning(request, ' | '.join(errors[:5]))
    return redirect('academics:campuses')


@login_required
def campus_list(request):
    return render(request, 'academics/campus_list.html', {'items': Campus.objects.all()})


@role_required(UserProfile.Role.ADMIN)
def campus_create(request):
    form = CampusForm(request.POST or None)
    if request.method == 'POST' and form.is_valid(): form.save(); messages.success(request, 'Campus saved.'); return redirect('academics:campuses')
    return render(request, 'academics/form.html', {'form': form, 'title': 'Add Campus'})


@role_required(UserProfile.Role.ADMIN)
def campus_edit(request, pk):
    campus = get_object_or_404(Campus, pk=pk); form = CampusForm(request.POST or None, instance=campus)
    if request.method == 'POST' and form.is_valid(): form.save(); messages.success(request, 'Campus updated.'); return redirect('academics:campuses')
    return render(request, 'academics/form.html', {'form': form, 'title': 'Edit Campus'})


@role_required(UserProfile.Role.ADMIN)
def campus_delete(request, pk):
    return _delete_setup_item(request, Campus, 'Campus', 'academics:campuses')


@role_required(UserProfile.Role.ADMIN)
def department_template_download(request):
    wb = Workbook(); ws = wb.active; ws.title = 'Departments'
    ws.append(DEPARTMENT_TEMPLATE_HEADERS); ws.append(['Main Campus', 'Department of Computer Science', 'TRUE']); ws.append(['Main Campus', 'Department of Physics', 'TRUE'])
    ws['E1'] = 'Active values'; ws['E2'] = 'TRUE'; ws['E3'] = 'FALSE'
    for col, width in {'A': 32, 'B': 42, 'C': 14, 'E': 16}.items(): ws.column_dimensions[col].width = width
    for cell in ws[1]: cell.font = Font(bold=True)
    return _xlsx_response(wb, 'department_upload_template.xlsx')


@role_required(UserProfile.Role.ADMIN)
def department_upload_excel(request):
    if request.method != 'POST': return redirect('academics:departments')
    form = DepartmentExcelUploadForm(request.POST, request.FILES)
    if not form.is_valid(): messages.error(request, 'Please select a valid Excel file.'); return redirect('academics:departments')
    try: sheet = load_workbook(form.cleaned_data['excel_file'], data_only=True).active
    except Exception: messages.error(request, 'The uploaded file could not be read. Please use the Excel template.'); return redirect('academics:departments')
    created = updated = skipped = 0; errors = []
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        campus_name = str(row[0] or '').strip() if len(row) > 0 else ''; name = str(row[1] or '').strip() if len(row) > 1 else ''; active_value = row[2] if len(row) > 2 else True
        if not campus_name and not name and not active_value: continue
        if not campus_name or not name: skipped += 1; errors.append(f'Row {index}: campus and department are required.'); continue
        campus, _ = Campus.objects.get_or_create(name=campus_name, defaults={'is_active': True})
        _, was_created = Department.objects.update_or_create(campus=campus, name=name, defaults={'is_active': _normalize_active(active_value)})
        created += int(was_created); updated += int(not was_created)
    messages.success(request, f'Excel upload complete. Created: {created}, Updated: {updated}, Skipped: {skipped}.')
    if errors: messages.warning(request, ' | '.join(errors[:5]))
    return redirect('academics:departments')


@login_required
def department_list(request):
    return render(request, 'academics/department_list.html', {'items': Department.objects.select_related('campus')})


@role_required(UserProfile.Role.ADMIN)
def department_create(request):
    form = DepartmentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid(): form.save(); messages.success(request, 'Department saved.'); return redirect('academics:departments')
    return render(request, 'academics/form.html', {'form': form, 'title': 'Add Department'})


@role_required(UserProfile.Role.ADMIN)
def department_edit(request, pk):
    department = get_object_or_404(Department, pk=pk); form = DepartmentForm(request.POST or None, instance=department)
    if request.method == 'POST' and form.is_valid(): form.save(); messages.success(request, 'Department updated.'); return redirect('academics:departments')
    return render(request, 'academics/form.html', {'form': form, 'title': 'Edit Department'})


@role_required(UserProfile.Role.ADMIN)
def department_delete(request, pk):
    return _delete_setup_item(request, Department, 'Department', 'academics:departments')


@login_required
def departments_by_campus(request):
    campus_id = request.GET.get('campus')
    qs = Department.objects.filter(is_active=True)
    if campus_id:
        qs = qs.filter(campus_id=campus_id)
    data = [{'id': d.pk, 'text': d.name} for d in qs.order_by('name')]
    return JsonResponse({'results': data})


@role_required(UserProfile.Role.ADMIN)
def program_template_download(request):
    wb = Workbook(); ws = wb.active; ws.title = 'Programs'
    ws.append(PROGRAM_TEMPLATE_HEADERS); ws.append(['BS Example Program', 'Bachelor', 'TRUE']); ws.append(['MS Example Program', 'Master', 'TRUE']); ws.append(['PhD Example Program', 'PhD', 'TRUE'])
    ws['E1'] = 'Allowed Levels'
    for row_no, (_, label) in enumerate(Program.Level.choices, start=2): ws[f'E{row_no}'] = label
    ws['G1'] = 'Active values'; ws['G2'] = 'TRUE'; ws['G3'] = 'FALSE'
    for col, width in {'A': 38, 'B': 18, 'C': 14, 'E': 18, 'G': 16}.items(): ws.column_dimensions[col].width = width
    for cell in ws[1]: cell.font = Font(bold=True)
    return _xlsx_response(wb, 'program_upload_template.xlsx')


@role_required(UserProfile.Role.ADMIN)
def program_upload_excel(request):
    if request.method != 'POST': return redirect('academics:programs')
    form = ProgramExcelUploadForm(request.POST, request.FILES)
    if not form.is_valid(): messages.error(request, 'Please select a valid Excel file.'); return redirect('academics:programs')
    try: sheet = load_workbook(form.cleaned_data['excel_file'], data_only=True).active
    except Exception: messages.error(request, 'The uploaded file could not be read. Please use the Excel template.'); return redirect('academics:programs')
    created = updated = skipped = 0; errors = []
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = str(row[0] or '').strip() if len(row) > 0 else ''; level_value = row[1] if len(row) > 1 else ''; active_value = row[2] if len(row) > 2 else True
        if not name and not level_value and not active_value: continue
        if not name: skipped += 1; errors.append(f'Row {index}: program name is missing.'); continue
        level = _normalize_program_level(level_value)
        if not level: skipped += 1; errors.append(f'Row {index}: invalid level "{level_value}".'); continue
        _, was_created = Program.objects.update_or_create(name=name, defaults={'level': level, 'is_active': _normalize_active(active_value)})
        created += int(was_created); updated += int(not was_created)
    messages.success(request, f'Excel upload complete. Created: {created}, Updated: {updated}, Skipped: {skipped}.')
    if errors: messages.warning(request, ' | '.join(errors[:5]))
    return redirect('academics:programs')


@login_required
def program_list(request): return render(request, 'academics/program_list.html', {'items': Program.objects.all()})
@role_required(UserProfile.Role.ADMIN)
def program_create(request):
    form = ProgramForm(request.POST or None)
    if request.method == 'POST' and form.is_valid(): form.save(); messages.success(request, 'Program saved.'); return redirect('academics:programs')
    return render(request, 'academics/form.html', {'form': form, 'title': 'Add Program'})
@role_required(UserProfile.Role.ADMIN)
def program_edit(request, pk):
    program = get_object_or_404(Program, pk=pk); form = ProgramForm(request.POST or None, instance=program)
    if request.method == 'POST' and form.is_valid(): form.save(); messages.success(request, 'Program updated.'); return redirect('academics:programs')
    return render(request, 'academics/form.html', {'form': form, 'title': 'Edit Program'})
@role_required(UserProfile.Role.ADMIN)
def program_delete(request, pk): return _delete_setup_item(request, Program, 'Program', 'academics:programs')

@login_required
def bank_list(request): return render(request, 'academics/bank_list.html', {'items': Bank.objects.all()})
@role_required(UserProfile.Role.ADMIN)
def bank_create(request):
    form = BankForm(request.POST or None)
    if request.method == 'POST' and form.is_valid(): form.save(); messages.success(request, 'Bank saved.'); return redirect('academics:banks')
    return render(request, 'academics/form.html', {'form': form, 'title': 'Add Bank'})
@role_required(UserProfile.Role.ADMIN)
def bank_edit(request, pk):
    bank = get_object_or_404(Bank, pk=pk); form = BankForm(request.POST or None, instance=bank)
    if request.method == 'POST' and form.is_valid(): form.save(); messages.success(request, 'Bank updated.'); return redirect('academics:banks')
    return render(request, 'academics/form.html', {'form': form, 'title': 'Edit Bank'})
@role_required(UserProfile.Role.ADMIN)
def bank_delete(request, pk): return _delete_setup_item(request, Bank, 'Bank', 'academics:banks')

@login_required
def courier_list(request): return render(request, 'academics/courier_list.html', {'items': CourierCompany.objects.all()})
@role_required(UserProfile.Role.ADMIN)
def courier_create(request):
    form = CourierCompanyForm(request.POST or None)
    if request.method == 'POST' and form.is_valid(): form.save(); messages.success(request, 'Courier company saved.'); return redirect('academics:couriers')
    return render(request, 'academics/form.html', {'form': form, 'title': 'Add Courier Company'})
@role_required(UserProfile.Role.ADMIN)
def courier_edit(request, pk):
    company = get_object_or_404(CourierCompany, pk=pk); form = CourierCompanyForm(request.POST or None, instance=company)
    if request.method == 'POST' and form.is_valid(): form.save(); messages.success(request, 'Courier company updated.'); return redirect('academics:couriers')
    return render(request, 'academics/form.html', {'form': form, 'title': 'Edit Courier Company'})
@role_required(UserProfile.Role.ADMIN)
def courier_delete(request, pk): return _delete_setup_item(request, CourierCompany, 'Courier Company', 'academics:couriers')
